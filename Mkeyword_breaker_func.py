import openpyxl

def generate_substrings(word):
    substrings = []
    for i in range(len(word)):
        for j in range(i + 1, len(word) + 1):
            substrings.append(word[i:j])
    return substrings

def append_substrings_to_cell(excel_file_path, sheet_name, row_number, column_number1):
    wb = openpyxl.load_workbook(excel_file_path)
    sheet = wb[sheet_name]

    column_number = ord("A") - ord('A') + 1
    cell = sheet.cell(row=row_number, column=column_number)
    data = cell.value
    if data is not None:  # Add this check
      substrings = generate_substrings(data)
      sheet.cell(row=row_number, column=column_number1).value = str(substrings)
      print(f"Data from cell {row_number}{column_number}: {data}")
    #substrings = generate_substrings(data)
      print(substrings)
    #sheet.cell(row=row_number, column=column_number1).value = str(substrings)
      print(f"writing data to cell {row_number}{column_number}: {substrings}")
    wb.save(excel_file_path)
    wb.close()

excel_file_path = "data_and_ids_g_use copy.xlsx"
sheet_name = "Version 2"
#row_number = 2
column_number1 = 5

#for i in from row 2 to row 1050 
for i in range(2, 1051):
  row_number = i
  append_substrings_to_cell(excel_file_path, sheet_name, row_number, column_number1)